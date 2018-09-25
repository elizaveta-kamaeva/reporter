from openpyxl.utils.cell import get_column_letter


def write_common(sheet, report_dict):
    title = 'Общие метрики'
    subtitles = ['', 'Сессии', 'Заказы', 'Выручка', 'Конверсия', 'RPS', 'AOV']
    row_names = ['Всего', 'Сессии без поиска', 'Сессии с поиском', 'С поиском / Всего']
    cells = {key: report_dict[key] for key in {'orders_total',
                                               'revenue_total',
                                               'sessions_total',
                                               'autocomplete_and_search_sessions_total',
                                               'autocomplete_and_search_sessions_revenue',
                                               'autocomplete_and_search_sessions_orders_total'}}
    # styling title
    sheet['A1'] = title
    sheet['A1'].style = 'title'
    sheet.merge_cells('A1:G1')

    # styling subtitle
    for s_row in range(2, 3):
        for s_col in range(1, 8):
            s_ind = get_column_letter(s_col) + str(s_row)
            sheet[s_ind] = subtitles[s_col-1]
            sheet[s_ind].style = 'subtitle'

    # styling percents
    for pc_row in range(3, 6):
        sheet['E'+str(pc_row)].number_format = '0.00%'
    for pc_col in range(2, 5):
        sheet[get_column_letter(pc_col)+'6'].number_format = '0.00%'

    # styling floats
    for fl_row in range(3, 6):
        for fl_col in range(6, 8):
            fl_ind = get_column_letter(fl_col) + str(fl_row)
            sheet[fl_ind].number_format = '#,##0.00'

    # writing row names
    for rn_row in range(3, 7):
        sheet['A' + str(rn_row)] = row_names[rn_row-3]

    # total
    sheet['B3'] = cells['sessions_total']
    sheet['C3'] = cells['orders_total']
    sheet['D3'] = cells['revenue_total']
    try:
        sheet['E3'] = round(sheet['C3'].value / sheet['B3'].value)
    except ZeroDivisionError:
        sheet['E3'] = 0
    try:
        sheet['F3'] = round(sheet['D3'].value / sheet['B3'].value, 2)
    except ZeroDivisionError:
        sheet['F3'] = 0
    try:
        sheet['G3'] = round(sheet['D3'].value / sheet['C3'].value, 2)
    except ZeroDivisionError:
        sheet['G3'] = 0

    # sessions without search
    sheet['B4'] = cells['sessions_total'] - cells['autocomplete_and_search_sessions_total']
    sheet['C4'] = cells['orders_total'] - cells['autocomplete_and_search_sessions_orders_total']
    sheet['D4'] = cells['revenue_total'] - cells['autocomplete_and_search_sessions_revenue']
    try:
        sheet['E4'] = round(sheet['C4'].value / sheet['B4'].value, 2)
    except ZeroDivisionError:
        sheet['E4'] = 0
    try:
        sheet['F4'] = round(sheet['D4'].value / sheet['B4'].value, 2)
    except ZeroDivisionError:
        sheet['F4'] = 0
    try:
        sheet['G4'] = round(sheet['D4'].value / sheet['C4'].value, 2)
    except ZeroDivisionError:
        sheet['G4'] = 0

    # sessions with search
    sheet['B5'] = cells['autocomplete_and_search_sessions_total']
    sheet['C5'] = cells['autocomplete_and_search_sessions_orders_total']
    sheet['D5'] = cells['autocomplete_and_search_sessions_revenue']
    try:
        sheet['E5'] = round(sheet['C5'].value / sheet['B5'].value, 2)
    except ZeroDivisionError:
        sheet['E5'] = 0
    try:
        sheet['F5'] = round(sheet['D5'].value / sheet['B5'].value, 2)
    except ZeroDivisionError:
        sheet['F5'] = 0
    try:
        sheet['G5'] = round(sheet['D5'].value / sheet['C5'].value, 2)
    except ZeroDivisionError:
        sheet['G5'] = 0

    # with search / total
    try:
        sheet['B6'] = round(cells['autocomplete_and_search_sessions_total'] / cells['sessions_total'], 2)
    except ZeroDivisionError:
        sheet['B6'] = 0
    try:
        sheet['C6'] = round(cells['autocomplete_and_search_sessions_orders_total'] / cells['orders_total'], 2)
    except ZeroDivisionError:
        sheet['C6'] = 0
    try:
        sheet['D6'] = round(cells['autocomplete_and_search_sessions_revenue'] / cells['revenue_total'], 2)
    except ZeroDivisionError:
        sheet['D6'] = 0
