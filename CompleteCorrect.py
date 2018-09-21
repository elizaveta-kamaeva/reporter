from openpyxl.utils.cell import get_column_letter


def write_acomplete(sheet, report_dict):
    title = 'Автоподсказки'
    subtitles = ['Сессии', 'Заказы', 'Выручка, руб.', 'CTR', 'Конверсия']
    cells = {key: report_dict[key] for key in {'autocomplete_sessions_total',
                                               'autocomplete_orders_total',
                                               'autocomplete_session_revenue',
                                               'autocomplete_ctr',
                                               'autocomplete_sessions_conversion'}}
    # styling title
    sheet['A9'] = title
    sheet['A9'].style = 'title'
    sheet.merge_cells('A9:E9')

    # styling subtitle
    for s_col in range(1, 6):
        s_ind = get_column_letter(s_col) + '10'
        sheet[s_ind].style = 'subtitle'
        sheet[s_ind] = subtitles[s_col-1]

    # styling centered text
    for s_col in range(1, 6):
        s_ind = get_column_letter(s_col) + '11'
        sheet[s_ind].style = 'centered'

    # styling percents
    sheet['D11'].number_format = '0.00%'
    sheet['E11'].number_format = '0.00%'

    # writing values
    sheet['A11'] = cells['autocomplete_sessions_total']
    sheet['B11'] = cells['autocomplete_orders_total']
    sheet['C11'] = cells['autocomplete_session_revenue']
    sheet['D11'] = cells['autocomplete_ctr'] / 100
    sheet['E11'] = cells['autocomplete_sessions_conversion'] / 100

    print('Autocomplete total stat done')


def write_clicks(sheet, report_dict):
    title = 'Клики по автоподказкам'
    row_names = ['Товары', 'Категории', 'Запросы', 'История', 'Всего']
    cells = {key: report_dict[key] for key in {'autocomplete_product_block_click',
                                               'autocomplete_category_block_click',
                                               'autocomplete_query_block_click',
                                               'autocomplete_history_block_click',
                                               'autocomplete_clicks'}}
    if cells['autocomplete_history_block_click'] == 0:
        del cells['autocomplete_history_block_click']
        del row_names[3]

    # styling title
    sheet['F9'] = title
    sheet['F9'].style = 'title'
    sheet.merge_cells('F9:G9')

    # styling percents
    sheet['D11'].number_format = '0.00%'
    sheet['E11'].number_format = '0.00%'

    # writing row names
    for rn_row in range(10, 10+len(cells)):
        sheet['F' + str(rn_row)] = row_names[rn_row-10]

    # writing values
    sheet['G10'] = cells['autocomplete_product_block_click']
    sheet['G11'] = cells['autocomplete_category_block_click']
    sheet['G12'] = cells['autocomplete_query_block_click']
    if 'autocomplete_history_block_click' in cells:
        sheet['G13'] = cells['autocomplete_history_block_click']
        sheet['G14'] = cells['autocomplete_clicks']
    else:
        sheet['G13'] = cells['autocomplete_clicks']

    print('Autocomplete clicks done')


def write_corrections(sheet, report_dict):
    title = 'Исправления'
    subtitles = ['Всего', 'Заказы', 'Выручка, руб.']
    cells = {key: report_dict[key] for key in {'aq_correction_total',
                                               'aq_correction_session_orders',
                                               'aq_correction_revenue'}}
    # styling title
    sheet['A15'] = title
    sheet['A15'].style = 'title'
    sheet.merge_cells('A15:C15')

    # styling subtitle
    for s_col in range(1, 4):
        s_ind = get_column_letter(s_col) + '16'
        sheet[s_ind].style = 'subtitle'
        sheet[s_ind] = subtitles[s_col-1]

    # styling centered text
    for s_col in range(1, 4):
        s_ind = get_column_letter(s_col) + '17'
        sheet[s_ind].style = 'centered'

    # writing values
    sheet['A17'] = cells['aq_correction_total']
    sheet['B17'] = cells['aq_correction_session_orders']
    sheet['C17'] = cells['aq_correction_revenue']

    print('Autocomplete clicks done')

