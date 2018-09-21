from openpyxl.utils.cell import get_column_letter
from collections import OrderedDict


def write_queries(sheet, report_dict):
    title = 'Поисковые запросы'
    row_names = ['Уникальные', 'Всего']
    cells = {key: report_dict[key] for key in {'unique_queries_total',
                                               'search_events_total'}}
    # styling title
    sheet['I1'] = title
    sheet['I1'].style = 'title'
    sheet.merge_cells('I1:J1')

    # writing row names
    for rn_row in range(2, 4):
        sheet['I' + str(rn_row)] = row_names[rn_row-2]

    # writing values
    sheet['J2'] = cells['unique_queries_total']
    sheet['J3'] = cells['search_events_total']

    print('Search queries done')


def write_popular(sheet, queries_dict):
    title = 'Популярные запросы'
    subtitles = ['Запрос', 'Количество']

    # styling title
    sheet['I7'] = title
    sheet['I7'].style = 'title'
    sheet.merge_cells('I7:J7')

    # styling subtitle
    for i in range(9, 11):
        sheet[get_column_letter(i) + '8'].style = 'subtitle'
        sheet[get_column_letter(i) + '8'] = subtitles[i-9]

    j = 0
    max_width = 0
    ordered_queries = OrderedDict(sorted(queries_dict.items(), key=lambda x: (-x[1], x[0])))
    for query in ordered_queries:
        sheet['I' + str(j+9)] = query
        sheet['J' + str(j+9)] = ordered_queries[query]
        if len(query) > max_width:
            max_width = len(query)
        j += 1
    sheet.column_dimensions['I'].width = max_width * 1.1

    print('Popular queries done')
