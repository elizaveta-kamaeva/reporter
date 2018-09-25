from os import startfile
from openpyxl import *
from openpyxl.styles import *
from openpyxl.utils.cell import get_column_letter

import CommonMetrics, CompleteCorrect, Queries


def set_path():
    path = 'C:\PyProjects\Reporter\\texts\\'
    print('The path to file is:', path)
    return path


def set_styles(wb):
    title_st = NamedStyle(name='title')
    title_st.font = Font(bold=True, name='Arial', size=11)
    title_st.alignment = Alignment(horizontal='center')
    title_st.fill = PatternFill('solid', fgColor='EFEFEF')
    wb.add_named_style(title_st)

    subtitle_st = NamedStyle(name='subtitle')
    subtitle_st.font = Font(name='Arial', size=11)
    subtitle_st.alignment = Alignment(horizontal='center')
    subtitle_st.fill = PatternFill('solid', fgColor='EFEFEF')
    wb.add_named_style(subtitle_st)

    centered_st = NamedStyle(name='centered')
    centered_st.font = Font(name='Arial', size=11)
    centered_st.alignment = Alignment(horizontal='center')
    centered_st.number_format = '#,##0'
    wb.add_named_style(centered_st)


def process(reports_array, date_array, filename):
    # creating workbook
    wb = Workbook()
    set_styles(wb)
    sheet = wb.active
    for i in range(len(date_array)):
        report_dict = reports_array[i]
        # as date_array looks like [<date_0>[<start>(year, month, day), <end>(year, month, day)], <date_1>[...]]
        start = '.'.join(str(j).rjust(2, '0') for j in (date_array[i][0][2], date_array[i][0][1]))
        end = '.'.join(str(j).rjust(2, '0') for j in (date_array[i][1][2], date_array[i][1][1]))
        sheet.title = '{}-{}'.format(start, end)

        # styling cells
        for s_row in range(1, 51):
            for s_col in range(1, 13):
                ind = get_column_letter(s_col) + str(s_row)
                sheet[ind].font = Font(name='Arial', size=11)
                sheet.column_dimensions[get_column_letter(s_col)].width = 15
                sheet.column_dimensions['A'].width = 25
                sheet[ind].number_format = '#,##0'

        # write metrics
        CommonMetrics.write_common(sheet, report_dict)
        CompleteCorrect.write_acomplete(sheet, report_dict)
        CompleteCorrect.write_clicks(sheet, report_dict)
        CompleteCorrect.write_corrections(sheet, report_dict)
        Queries.write_queries(sheet, report_dict)
        Queries.write_popular(sheet, report_dict['top_search_queries'])
        print('Data loaded into virtual table for {} - {}'.format(start, end))

        if i+1 < len(date_array):
            sheet = wb.create_sheet()

    path = set_path() + filename
    while True:
        try:
            wb.save(path)
            break
        except PermissionError:
            input('I can\'t write to an opened file. Close it, please, and hit ENTER.')
    startfile(path)
